#!/usr/bin/env python3
"""Cloud Cost Optimizer API — FastAPI server for automated cloud cost analysis.

Features:
- User authentication (login/register with session tokens)
- AWS CUR upload & analysis
- AWS CLI direct billing scan
- Multi-user support with isolated analysis results
"""

import hashlib
import hmac
import json
import secrets
import sys
import tempfile
import time
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent.parent / "cli"))
from optimizer import CostAnalyzer

from fastapi import Depends, FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse, Response

app = FastAPI(title="Cloud Cost Optimizer API", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"])


# ── Rate Limiting ────────────────────────────────────────────────────

# Simple in-memory rate limiter (MVP — use Redis for production)
_RATE_LIMIT: dict = {}  # ip -> [(timestamp, ...)]
RATE_LIMIT_MAX = 60  # requests per minute
RATE_LIMIT_WINDOW = 60  # seconds


async def rate_limit_middleware(request: Request, call_next):
    """Simple in-memory rate limiter per IP."""
    import time as _time
    client_ip = request.client.host if request.client else "unknown"
    now = _time.time()

    # Skip rate limiting for health checks
    if request.url.path == "/api/health":
        return await call_next(request)

    # Clean old entries
    if client_ip in _RATE_LIMIT:
        _RATE_LIMIT[client_ip] = [
            t for t in _RATE_LIMIT[client_ip] if now - t < RATE_LIMIT_WINDOW
        ]
    else:
        _RATE_LIMIT[client_ip] = []

    # Check limit
    if len(_RATE_LIMIT[client_ip]) >= RATE_LIMIT_MAX:
        return JSONResponse(
            status_code=429,
            content={"detail": "Rate limit exceeded. Max 60 requests per minute."},
        )

    _RATE_LIMIT[client_ip].append(now)
    response = await call_next(request)
    return response


app.middleware("http")(rate_limit_middleware)

# ── User Authentication ──────────────────────────────────────────────

# In-memory user store (MVP — use database for production)
USERS_DB: dict = {}  # username -> {password_hash, salt, created}
TOKENS_DB: dict = {}  # token -> {username, created, user_id}
USER_ANALYSIS: dict = {}  # user_id -> list of analysis results
NEXT_USER_ID = 1


def _hash_password(password: str, salt: str) -> str:
    """Simple SHA256 salted hash (MVP only — use bcrypt for production)"""
    return hashlib.sha256(f"{salt}{password}".encode()).hexdigest()


def create_user(username: str, password: str) -> dict:
    """Create a new user account."""
    if username in USERS_DB:
        raise ValueError("Username already exists")
    salt = secrets.token_hex(16)
    global NEXT_USER_ID
    user_id = NEXT_USER_ID
    NEXT_USER_ID += 1
    USERS_DB[username] = {
        "password_hash": _hash_password(password, salt),
        "salt": salt,
        "created": datetime.now().isoformat(),
        "user_id": user_id,
    }
    USER_ANALYSIS[user_id] = []
    return {"user_id": user_id, "username": username}


def authenticate_user(username: str, password: str) -> str:
    """Authenticate user and return session token."""
    if username not in USERS_DB:
        return None
    user = USERS_DB[username]
    if _hash_password(password, user["salt"]) != user["password_hash"]:
        return None
    token = secrets.token_urlsafe(32)
    user_id = get_user_id(username)
    # Ensure USER_ANALYSIS has entry for this user
    if user_id not in USER_ANALYSIS:
        USER_ANALYSIS[user_id] = []
    TOKENS_DB[token] = {
        "username": username,
        "created": time.time(),
        "user_id": user_id,
    }
    return token


def get_user_id(username: str) -> int:
    """Get user_id from username."""
    if username in USERS_DB:
        return USERS_DB[username].get("user_id", 1)
    return 1


def get_user_by_token(token: Optional[str] = None) -> dict:
    """Validate token and return user info."""
    if not token or token not in TOKENS_DB:
        raise HTTPException(status_code=401, detail="Invalid or missing token")
    return TOKENS_DB[token]


def require_auth(authorization: Optional[str] = Header(None)) -> dict:
    """Dependency: require valid auth token."""
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization header required")
    token = authorization.replace("Bearer ", "") if authorization.startswith("Bearer ") else authorization
    return get_user_by_token(token)


# ── Create default admin on startup ──────────────────────────────────

try:
    create_user("admin", "admin123")
except ValueError:
    pass  # Already exists


# ── Public Endpoints (no auth required) ──────────────────────────────

analysis_results: list[dict] = []  # Legacy: kept for backward compat


@app.get("/", response_class=HTMLResponse)
async def dashboard():
    return _DASHBOARD_HTML


@app.get("/api/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


@app.get("/api/metrics")
async def metrics():
    """Prometheus-style metrics endpoint (no auth required for monitoring)."""
    total_users = len(USERS_DB)
    total_tokens = len(TOKENS_DB)
    total_analyses = sum(len(v) for v in USER_ANALYSIS.values())
    return {
        "users_total": total_users,
        "active_sessions": total_tokens,
        "analyses_total": total_analyses,
        "timestamp": datetime.now().isoformat(),
    }


@app.post("/api/register")
async def register(username: str, password: str):
    """Register a new user."""
    if len(username) < 3:
        raise HTTPException(status_code=400, detail="Username too short (min 3 chars)")
    if len(password) < 4:
        raise HTTPException(status_code=400, detail="Password too short (min 4 chars)")
    try:
        result = create_user(username, password)
        return {"message": "User created", **result}
    except ValueError as e:
        raise HTTPException(status_code=409, detail=str(e))


@app.post("/api/login")
async def login(username: str, password: str):
    """Login and get session token."""
    token = authenticate_user(username, password)
    if not token:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    user = TOKENS_DB[token]
    return {
        "token": token,
        "username": user["username"],
        "user_id": user["user_id"],
        "message": "Login successful. Use this token in Authorization: Bearer <token>",
    }


# ── Protected Endpoints (auth required) ──────────────────────────────


@app.post("/api/analyze")
async def analyze(file: UploadFile = File(...), user: dict = Depends(require_auth)):
    try:
        contents = await file.read()
        analyzer = CostAnalyzer()

        if file.filename.endswith('.gz'):
            import gzip
            text = gzip.decompress(contents).decode('utf-8')
        else:
            text = contents.decode('utf-8')

        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
            f.write(text)
            temp_path = f.name

        resources = analyzer.parse_aws_cur(temp_path)
        Path(temp_path).unlink()

        recs = analyzer.generate_recommendations()
        idle = [r for r in resources if r.is_idle]
        total_savings = sum(r.estimated_savings_monthly for r in recs)

        result = {
            "resources_analyzed": len(resources),
            "idle_resources": len(idle),
            "recommendations": [asdict(r) for r in recs],
            "total_savings": total_savings,
            "generated": datetime.now().isoformat(),
        }

        USER_ANALYSIS[user["user_id"]].append(result)
        return result

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/recommendations")
async def get_recommendations(user: dict = Depends(require_auth)):
    analyses = USER_ANALYSIS.get(user["user_id"], [])
    if not analyses:
        return {"message": "No analyses run yet. POST /api/analyze first."}
    return analyses[-1]


@app.get("/api/report")
async def get_report(format: str = "markdown", user: dict = Depends(require_auth)):
    analyses = USER_ANALYSIS.get(user["user_id"], [])
    if not analyses:
        return {"message": "No analyses run yet."}
    result = analyses[-1]
    if format == "json":
        return JSONResponse(content=result)
    else:
        lines = [
            "# Cloud Cost Optimization Report",
            f"**Generated**: {result['generated']}",
            f"**Resources**: {result['resources_analyzed']}",
            f"**Idle**: {result['idle_resources']}",
            "",
            f"## Total Potential Savings: ${result['total_savings']:,.2f}/month",
            "",
        ]
        for r in result['recommendations'][:20]:
            lines.append(
                f"- [{r['confidence'].upper()}] {r['resource_type']} {r['resource_id']}: "
                f"{r['action']} -> ${r['estimated_savings_monthly']:,.2f}/mo"
            )
        return HTMLResponse(content="<br>".join(lines))


@app.post("/api/aws-scan")
async def aws_scan(
    profile: Optional[str] = None,
    region: str = "us-east-1",
    user: dict = Depends(require_auth)
):
    """Scan AWS directly for idle resources and cost optimization."""
    try:
        from cli.aws_cli import AWSConnector

        conn = AWSConnector(profile=profile, region=region)
        idle = conn.full_scan()

        # Convert to recommendation format
        recs = []
        for r in idle:
            recs.append({
                "resource_id": r.resource_id,
                "resource_type": r.resource_type,
                "action": "terminate" if r.resource_type in ("EC2", "EBS") else "review",
                "estimated_savings_monthly": r.cost_30d,
                "confidence": "high" if r.is_idle else "medium",
                "description": r.idle_reason,
                "implementation_steps": [],
            })

        result = {
            "scan_type": "aws_direct",
            "profile": profile or "default",
            "region": region,
            "resources_scanned": len(idle),
            "idle_resources": len([r for r in idle if r.is_idle]),
            "recommendations": recs,
            "total_savings": sum(r.cost_30d for r in idle),
            "generated": datetime.now().isoformat(),
        }

        USER_ANALYSIS[user["user_id"]].append(result)
        return result

    except ImportError:
        raise HTTPException(status_code=500, detail="boto3 not installed. Run: pip install boto3")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/export-csv")
async def export_csv(user: dict = Depends(require_auth)):
    """Export recommendations as CSV for download."""
    import csv
    import io
    analyses = USER_ANALYSIS.get(user["user_id"], [])
    if not analyses:
        raise HTTPException(status_code=404, detail="No analyses to export")
    result = analyses[-1]
    if not result.get("recommendations"):
        raise HTTPException(status_code=404, detail="No recommendations to export")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Resource ID", "Resource Type", "Action", "Monthly Savings", "Confidence", "Description"])
    for r in result["recommendations"]:
        writer.writerow([
            r["resource_id"],
            r["resource_type"],
            r["action"],
            f"${r['estimated_savings_monthly']:.2f}",
            r["confidence"],
            r.get("description", ""),
        ])
    csv_content = output.getvalue()
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=cost-optimizer-report.csv"},
    )


@app.get("/api/export-excel")
async def export_excel(user: dict = Depends(require_auth)):
    """Export recommendations as Excel file for download."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        analyses = USER_ANALYSIS.get(user["user_id"], [])
        if not analyses:
            raise HTTPException(status_code=404, detail="No analyses to export")
        result = analyses[-1]
        if not result.get("recommendations"):
            raise HTTPException(status_code=404, detail="No recommendations to export")

        wb = Workbook()
        ws = wb.active
        ws.title = "Recommendations"

        # Header styling
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Summary sheet
        ws.cell(row=1, column=1, value="Cloud Cost Optimization Report")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws.cell(row=2, column=1, value=f"Generated: {result['generated']}")
        ws.cell(row=3, column=1, value=f"Resources Analyzed: {result['resources_analyzed']}")
        ws.cell(row=4, column=1, value=f"Idle Resources: {result['idle_resources']}")
        ws.cell(row=5, column=1, value=f"Total Monthly Savings: ${result['total_savings']:,.2f}")

        # Recommendations sheet
        ws2 = wb.create_sheet("Recommendations")
        headers = ["#", "Resource ID", "Type", "Action", "Monthly Savings", "Confidence", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, rec in enumerate(result["recommendations"], 1):
            ws2.cell(row=i+1, column=1, value=i)
            ws2.cell(row=i+1, column=2, value=rec["resource_id"])
            ws2.cell(row=i+1, column=3, value=rec["resource_type"])
            ws2.cell(row=i+1, column=4, value=rec["action"])
            ws2.cell(row=i+1, column=5, value=rec["estimated_savings_monthly"])
            ws2.cell(row=i+1, column=6, value=rec["confidence"])
            ws2.cell(row=i+1, column=7, value=rec.get("description", ""))

        # Auto-fit column widths
        for col in ws2.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws2.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

        # Save to BytesIO
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=cost-optimizer-report.xlsx"},
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")


@app.get("/api/history")
async def get_history(user: dict = Depends(require_auth)):
    """Get full analysis history."""
    analyses = USER_ANALYSIS.get(user["user_id"], [])
    return {
        "count": len(analyses),
        "analyses": [
            {
                "id": i + 1,
                "generated": a["generated"],
                "resources": a["resources_analyzed"],
                "idle": a["idle_resources"],
                "recommendations": len(a.get("recommendations", [])),
                "savings": a["total_savings"],
            }
            for i, a in enumerate(analyses)
        ],
    }


@app.get("/api/user")
async def get_user_info(user: dict = Depends(require_auth)):
    """Get current user info and analysis history."""
    analyses = USER_ANALYSIS.get(user["user_id"], [])
    return {
        "username": user["username"],
        "user_id": user["user_id"],
        "analyses_count": len(analyses),
        "last_analysis": analyses[-1]["generated"] if analyses else None,
    }


_DASHBOARD_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Cloud Cost Optimizer</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f5f5f5}
.hdr{background:linear-gradient(135deg,#1a73e8,#0d47a1);color:#fff;padding:20px 30px;box-shadow:0 2px 4px rgba(0,0,0,.1)}
.hdr h1{font-size:1.5em;margin-bottom:4px}.hdr p{opacity:.8;font-size:.9em}
.cnt{max-width:1000px;margin:20px auto;padding:0 20px}
.card{background:#fff;border-radius:12px;padding:24px;margin-bottom:20px;box-shadow:0 1px 3px rgba(0,0,0,.1)}
.card h2{font-size:1.2em;margin-bottom:16px;color:#333}
.ua{border:2px dashed #ccc;border-radius:8px;padding:40px;text-align:center;cursor:pointer;transition:all .2s}
.ua:hover{border-color:#1a73e8;background:#f8f9ff}.ua.dragover{border-color:#1a73e8;background:#e8f0fe}
.ua input{display:none}.ua .ico{font-size:3em;margin-bottom:10px}
.btn{background:#1a73e8;color:#fff;padding:10px 24px;border:none;border-radius:6px;cursor:pointer;font-size:1em;transition:background .2s}
.btn:hover{background:#1557b0}.btn:disabled{background:#ccc;cursor:not-allowed}
.btn2{background:#34a853}.btn2:hover{background:#2d8f47}
.metrics{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:20px}
.mc{background:#f8f9fa;padding:20px;border-radius:8px;text-align:center}
.mv{font-size:2em;font-weight:700;color:#1a73e8}.ml{font-size:.85em;color:#666;margin-top:4px}
.mv.sv{color:#34a853}.mv.wn{color:#f9ab00}.mv.dn{color:#ea4335}
table{width:100%;border-collapse:collapse;margin-top:16px}
th{background:#f8f9fa;padding:12px;text-align:left;font-size:.85em;color:#666;border-bottom:2px solid #e0e0e0}
td{padding:12px;border-bottom:1px solid #f0f0f0}tr:hover{background:#f8f9ff}
.badge{display:inline-block;padding:2px 8px;border-radius:12px;font-size:.75em;font-weight:600}
.bh{background:#fce8e6;color:#c5221d}.bm{background:#fef7e0;color:#ea8600}.bl{background:#e6f4ea;color:#137333}
.loading{text-align:center;padding:40px;color:#666}
.spinner{border:3px solid #f3f3f3;border-top:3px solid #1a73e8;border-radius:50%;width:40px;height:40px;animation:spin 1s linear infinite;margin:0 auto 10px}
@keyframes spin{0%{transform:rotate(0)}100%{transform:rotate(360deg)}}
.error{background:#fce8e6;color:#c5221d;padding:12px;border-radius:8px;margin-top:10px}
.success{background:#e6f4ea;color:#137333;padding:12px;border-radius:8px;margin-top:10px}
.actions{display:flex;gap:10px;margin-top:16px}.hidden{display:none}
.pricing-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:20px;margin-top:16px}
.p-tier{text-align:center;padding:28px 20px;border-radius:12px;border:2px solid #e0e0e0;background:#fff;transition:all .2s;position:relative}
.p-tier:hover{border-color:#1a73e8;box-shadow:0 4px 12px rgba(26,115,232,.15)}
.p-tier.featured{border-color:#1a73e8;background:linear-gradient(180deg,#f8f9ff,#fff)}
.p-tier.featured:hover{box-shadow:0 4px 20px rgba(26,115,232,.25)}
.p-tier h3{font-size:1.1em;color:#333;margin-bottom:8px}
.p-price{font-size:2.2em;font-weight:700;color:#1a73e8;margin:12px 0 4px}
.p-price span{font-size:.45em;color:#666;font-weight:400}
.p-features{list-style:none;padding:0;margin:16px 0 0;text-align:left}
.p-features li{padding:6px 0;font-size:.9em;color:#555;border-bottom:1px solid #f5f5f5}
.p-features li:last-child{border-bottom:none}
.p-features li::before{content:"&#10003; ";color:#34a853;font-weight:700}
.p-badge{position:absolute;top:-12px;right:16px;background:#f9ab00;color:#fff;padding:3px 12px;border-radius:12px;font-size:.72em;font-weight:700;text-transform:uppercase;letter-spacing:.5px}
.p-badge.active{background:#34a853}
#costChart{width:100%;height:300px;background:#f8f9fa;border-radius:8px}
</style>
</head>
<body>
<div class="hdr"><h1>Cloud Cost Optimizer</h1><p>Upload your AWS CUR to identify waste and optimize costs</p></div>
<div class="cnt">
<!-- Auth Panel -->
<div class="card" id="authPanel">
<h2>🔐 Login / Register</h2>
<div style="display:flex;gap:20px;flex-wrap:wrap">
<div style="flex:1;min-width:250px">
<input type="text" id="authUser" placeholder="Username" style="width:100%;padding:10px;border:1px solid #ddd;border-radius:6px;margin-bottom:10px;font-size:1em">
<input type="password" id="authPass" placeholder="Password" style="width:100%;padding:10px;border:1px solid #ddd;border-radius:6px;margin-bottom:10px;font-size:1em">
<button class="btn" onclick="doLogin()" style="width:100%;margin-bottom:8px">Login</button>
<button class="btn btn2" onclick="doRegister()" style="width:100%">Register</button>
<div id="authMsg" style="margin-top:10px"></div>
</div>
<div style="flex:1;min-width:250px;background:#f8f9fa;padding:16px;border-radius:8px">
<h3 style="margin-bottom:8px">Why Login?</h3>
<ul style="padding-left:20px;line-height:1.8;color:#555">
<li>Upload and analyze AWS CUR files</li>
<li>Export recommendations as CSV/Excel</li>
<li>Track analysis history</li>
<li>AWS direct billing scan</li>
</ul>
</div>
</div>
</div>
<!-- Main App (hidden until login) -->
<div id="mainApp" style="display:none">
<div class="card">
<h2>Upload AWS CUR File</h2>
<div class="ua" id="ua" onclick="document.getElementById('f').click()">
<div class="ico">&#128230;</div><p>Drag & drop CSV here, or click to browse</p>
<p style="font-size:.85em;color:#999;margin-top:8px">Supports .csv and .csv.gz</p>
<input type="file" id="f" accept=".csv,.csv.gz">
</div>
<div id="fn" style="margin-top:10px;color:#666"></div>
<div class="actions">
<button class="btn" id="ab" onclick="run()" disabled>Analyze Costs</button>
<button class="btn btn2 hidden" id="db" onclick="dl()">Download Report</button>
<button class="btn btn2 hidden" id="ecb" onclick="exportCsv()">Export CSV</button>
<button class="btn hidden" id="hb" onclick="showHistory()">History</button>
</div>
<div id="msg"></div>
</div>
<div id="res" class="hidden">
<div class="metrics" id="mt"></div>
<div class="card">
<h2>Cost Trend</h2>
<canvas id="costChart" style="max-height:300px"></canvas>
</div>
<div class="card"><h2>Optimization Recommendations</h2><div id="rc"></div></div>
</div>
<div class="card">
<h2>Pricing</h2>
<div class="pricing-grid">
<div class="p-tier">
<span class="p-badge active">Available</span>
<h3>Free</h3>
<div class="p-price">$0<span>/mo</span></div>
<ul class="p-features">
<li>1 user</li>
<li>3 analyses per month</li>
<li>CSV upload</li>
</ul>
</div>
<div class="p-tier featured">
<span class="p-badge">Coming Soon</span>
<h3>Pro</h3>
<div class="p-price">$9<span>/mo</span></div>
<ul class="p-features">
<li>5 users</li>
<li>Unlimited analyses</li>
<li>AWS direct scan</li>
</ul>
</div>
<div class="p-tier">
<span class="p-badge">Coming Soon</span>
<h3>Team</h3>
<div class="p-price">$29<span>/mo</span></div>
<ul class="p-features">
<li>Unlimited users</li>
<li>Multi-region</li>
<li>Multi-account</li>
<li>API access</li>
</ul>
</div>
</div>
</div>
<div class="card">
<h2>How It Works</h2>
<ol style="padding-left:20px;line-height:1.8">
<li>Export AWS Cost & Usage Report from Billing Console</li>
<li>Upload the CSV file above</li>
<li>Get instant analysis of idle resources and cost optimization recommendations</li>
<li>Follow actionable steps to reduce cloud costs</li>
</ol>
</div>
</div>
</div>
<!-- End mainApp -->
<!-- User Bar (shown after login) -->
<div id="userBar" class="card" style="display:none;padding:12px 24px">
<div style="display:flex;justify-content:space-between;align-items:center">
<div>
<strong id="userName"></strong>
<span style="color:#666;margin-left:12px;font-size:.9em" id="userStats"></span>
</div>
<button class="btn" style="background:#ea4335;padding:8px 16px;font-size:.9em" onclick="doLogout()">Logout</button>
</div>
</div>
<script>
const $=id=>document.getElementById(id);
// Check for saved token on load
(function(){const t=localStorage.getItem('cco_token');if(t){window._token=t;$('#authPanel').style.display='none';$('#mainApp').style.display='block';loadUserInfo()}})();
function doLogin(){const u=$('authUser').value,p=$('authPass').value,msg=$('authMsg');if(!u||!p){msg.className='error';msg.textContent='Please enter username and password';return}fetch('/api/login',{method:'POST',headers:{'Content-Type':'application/x-www-form-urlencoded'},body:`username=${encodeURIComponent(u)}&password=${encodeURIComponent(p)}`}).then(r=>r.json()).then(d=>{if(d.token){window._token=d.token;localStorage.setItem('cco_token',d.token);$('#authPanel').style.display='none';$('#mainApp').style.display='block';loadUserInfo();msg.className='success';msg.textContent='Logged in as '+d.username}else{msg.className='error';msg.textContent=d.detail||'Login failed'}}).catch(e=>{msg.className='error';msg.textContent='Error: '+e.message})}
function doRegister(){const u=$('authUser').value,p=$('authPass').value,msg=$('authMsg');if(!u||!p){msg.className='error';msg.textContent='Please enter username and password';return}fetch('/api/register',{method:'POST',headers:{'Content-Type':'application/x-www-form-urlencoded'},body:`username=${encodeURIComponent(u)}&password=${encodeURIComponent(p)}`}).then(r=>r.json()).then(d=>{if(d.user_id){msg.className='success';msg.textContent='Account created! Now click Login'}else{msg.className='error';msg.textContent=d.detail||'Registration failed'}}).catch(e=>{msg.className='error';msg.textContent='Error: '+e.message})}
function doLogout(){localStorage.removeItem('cco_token');window._token=null;location.reload()}
function loadUserInfo(){if(!window._token)return;fetch('/api/user',{headers:{'Authorization':'Bearer '+window._token}}).then(r=>r.json()).then(d=>{$('userName').textContent='👤 '+d.username;$('userStats').textContent=d.analyses_count+' analyses';$('userBar').style.display='block'}).catch(()=>{})}
const f=$('f'),ab=$('ab'),ua=$('ua'),fn=$('fn'),msg=$('msg'),res=$('res');
f.addEventListener('change',()=>{if(f.files.length){ab.disabled=false;fn.textContent='Selected: '+f.files[0].name}});
ua.addEventListener('dragover',e=>{e.preventDefault();ua.classList.add('dragover')});
ua.addEventListener('dragleave',()=>ua.classList.remove('dragover'));
ua.addEventListener('drop',e=>{e.preventDefault();ua.classList.remove('dragover');if(e.dataTransfer.files.length){f.files=e.dataTransfer.files;ab.disabled=false;fn.textContent='Selected: '+e.dataTransfer.files[0].name}});
async function run(){
if(!f.files[0])return;
ab.disabled=true;ab.textContent='Analyzing...';
msg.className='loading';msg.innerHTML='<div class="spinner"></div><p>Processing...</p>';
res.classList.add('hidden');$('db').classList.add('hidden');
const fd=new FormData();fd.append('file',f.files[0]);
try{
const r=await fetch('/api/analyze',{method:'POST',body:fd});
const d=await r.json();
if(r.ok){show('Analysis complete! '+(d.recommendations.length||0)+' opportunities found.','success');showRes(d);$('db').classList.remove('hidden');$('ecb').classList.remove('hidden');$('hb').classList.remove('hidden');window._lastRes=d}
else show('Error: '+(d.error||'Analysis failed'),'error');
}catch(e){show('Network error: '+e.message,'error')}
ab.disabled=false;ab.textContent='Analyze Costs';
}
function show(t,c){msg.className=c;msg.textContent=t}
function exportCsv(){$('ecb').textContent='Exporting...';fetch('/api/export-csv',{headers:{'Authorization':'Bearer '+window._token}}).then(r=>r.blob()).then(b=>{const u=URL.createObjectURL(b),a=document.createElement('a');a.href=u;a.download='cost-optimizer-report.csv';a.click();URL.revokeObjectURL(u);$('ecb').textContent='Export CSV'}).catch(()=>{$('ecb').textContent='Export CSV'})}
function showHistory(){fetch('/api/history',{headers:{'Authorization':'Bearer '+window._token}}).then(r=>r.json()).then(d=>{if(!d.count){show('No analysis history','error');return}let h='<table><tr><th>#</th><th>Date</th><th>Resources</th><th>Idle</th><th>Recs</th><th>Savings</th></tr>';d.analyses.reverse().forEach(a=>{h+=`<tr><td>${a.id}</td><td>${a.generated}</td><td>${a.resources}</td><td>${a.idle}</td><td>${a.recommendations}</td><td>$${a.savings.toFixed(2)}</td></tr>`});h+='</table>';const card=document.createElement('div');card.className='card';card.id='historyCard';card.innerHTML=`<h2>Analysis History (${d.count})</h2>`+h;const pricing=document.querySelector('.pricing-grid')?.closest('.card');if(pricing)pricing.parentNode.insertBefore(card,pricing);else res.parentNode.appendChild(card);show('History loaded','success')}).catch(e=>show('Error: '+e.message,'error'))}
function showRes(d){
res.classList.remove('hidden');
$('mt').innerHTML=`<div class="mc"><div class="mv">${d.resources_analyzed}</div><div class="ml">Resources</div></div>
<div class="mc"><div class="mv wn">${d.idle_resources}</div><div class="ml">Idle</div></div>
<div class="mc"><div class="mv sv">${d.recommendations.length}</div><div class="ml">Recommendations</div></div>
<div class="mc"><div class="mv sv">$${d.total_savings.toLocaleString(undefined,{minimumFractionDigits:2,maximumFractionDigits:2})}</div><div class="ml">Monthly Savings</div></div>`;
drawChart(d.recommendations);
const rc=$('rc');
if(!d.recommendations.length){rc.innerHTML='<p style="color:#666">No recommendations. Your resources look optimized!</p>';return}
let h='<table><tr><th>Resource</th><th>Type</th><th>Action</th><th>Monthly Savings</th><th>Confidence</th></tr>';
for(const r of d.recommendations.slice(0,20)){
const bc=r.confidence==='high'?'bh':r.confidence==='medium'?'bm':'bl';
h+=`<tr><td class="rid">${r.resource_id}</td><td>${r.resource_type}</td><td>${r.action}</td><td>$${r.estimated_savings_monthly.toLocaleString(undefined,{minimumFractionDigits:2})}</td><td><span class="badge ${bc}">${r.confidence}</span></td></tr>`;
}
h+='</table>';rc.innerHTML=h;
}
function dl(){
const d=$('res').classList.contains('hidden');
if(d){show('No analysis to download','error');return}
window.location.href='/api/report?format=markdown';
}
function drawChart(recs){
const canvas=$('costChart');
if(!canvas||!recs.length)return;
const ctx=canvas.getContext('2d');
const W=canvas.width=canvas.offsetWidth;
const H=canvas.height=300;
ctx.clearRect(0,0,W,H);
const top=recs.slice(0,8).sort((a,b)=>b.estimated_savings_monthly-a.estimated_savings_monthly);
if(!top.length)return;
const maxVal=Math.max(...top.map(r=>r.estimated_savings_monthly));
const barW=(W-80)/top.length;
const chartH=H-60;
ctx.fillStyle='#666';ctx.font='12px sans-serif';
ctx.fillText('Monthly Savings by Resource',20,20);
ctx.strokeStyle='#e0e0e0';ctx.beginPath();ctx.moveTo(40,30);ctx.lineTo(40,H-30);ctx.lineTo(W-20,H-30);ctx.stroke();
top.forEach((r,i)=>{
const x=50+i*barW;
const barH=(r.estimated_savings_monthly/maxVal)*chartH;
const y=H-30-barH;
const grad=ctx.createLinearGradient(x,y,x,H-30);
grad.addColorStop(0,'#1a73e8');grad.addColorStop(1,'#34a853');
ctx.fillStyle=grad;ctx.fillRect(x,y,barW-10,barH);
ctx.fillStyle='#333';ctx.font='10px sans-serif';
const label=r.resource_id.length>12?r.resource_id.slice(0,12)+'...':r.resource_id;
ctx.save();ctx.translate(x+barW/2-5,H-15);ctx.rotate(-0.5);ctx.fillText(label,0,0);ctx.restore();
ctx.fillStyle='#1a73e8';ctx.font='11px sans-serif';
ctx.fillText('$'+r.estimated_savings_monthly.toFixed(0),x+5,y-5);
});
}
<\/script>
</body></html>"""


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8765)
